import os
import time
import shutil
from datetime import datetime
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
import openpyxl
import pandas as pd
from datetime import datetime, timedelta
def wait_for_downloads(directory, timeout=30):
    seconds = 0
    dl_wait = True
    while dl_wait and seconds < timeout:
        time.sleep(1)
        dl_wait = False
        for fname in os.listdir(directory):
            if fname.endswith('.crdownload'):
                dl_wait = True
        seconds += 1
    return not dl_wait

def click_checkbox(driver, checkbox):
    try:
        if not checkbox.is_selected():
            driver.execute_script("arguments[0].click();", checkbox)
            print("Checkbox clicked")
        else:
            print("Checkbox is already checked")
    except StaleElementReferenceException:
        print("Checkbox element is stale, re-finding...")
        checkboxes = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, "//input[@type='checkbox']")))
        for checkbox in checkboxes:
            click_checkbox(driver, checkbox)

def initialize_environment():
    # file_path = "try1.xlsx"
    file_path = "RO_Main.xlsx"
    download_path = Path.home() / "Downloads"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["Sheet1"]
    row_count = sheet.max_row

    chrome_options = Options()
    chrome_options.add_argument("--remote-allow-origins=*")
    prefs = {"download.default_directory": str(download_path)}
    chrome_options.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(options=chrome_options)
    driver.maximize_window()

    driver.get("https://cs.rowriter.com/")
    print(driver.title)

    driver.implicitly_wait(40)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "loginLink"))).click()

    return driver, sheet, row_count, download_path

def setup_logging(log_file_path):
    if not os.path.exists(log_file_path):
        df = pd.DataFrame(columns=["Clientname", "Username", "Password", "Report", "Subreport", "Frequency", "Status", "Date", "Time", "Additional Info"])
        df.to_excel(log_file_path, index=False, engine='openpyxl')

def log_operation(log_file_path, clientname, username, password, report, subreport, frequency, status, additional_info=""):
    df = pd.read_excel(log_file_path, engine='openpyxl') if os.path.exists(log_file_path) else pd.DataFrame(columns=["Date", "Time", "Clientname", "Username", "Password", "Report", "Subreport", "Frequency", "Status", "Additional Info"])

    now = datetime.now()
    today_date = now.strftime('%d/%m/%Y')
    current_time = now.strftime('%H:%M:%S')
    new_entry = {
        "Date": today_date,
        "Time": current_time,
        "Clientname": clientname,
        "Username": username,
        "Password": password,
        "Report": report,
        "Subreport": subreport,
        "Frequency": frequency,
        "Status": status,
        "Additional Info": additional_info
    }
    df = df.append(new_entry, ignore_index=True)
    df.to_excel(log_file_path, index=False, engine='openpyxl')
    
def create_folder_structure(destination_dir, clientname, year, month, day, category_folder, report_name):
    user_folder = os.path.join(destination_dir, clientname)
    year_folder = os.path.join(user_folder, year)
    month_folder = os.path.join(year_folder, month)
    day_folder = os.path.join(month_folder, day)
    category_folder_path = os.path.join(day_folder, category_folder)
    report_folder_path = os.path.join(category_folder_path, report_name)
    
    os.makedirs(report_folder_path, exist_ok=True)

    return report_folder_path
 


def process_rows(driver, sheet, row_count, download_path):
    destination_dir = os.path.expanduser("final")
    log_file_path = os.path.join(destination_dir, "file_transfer_log.xlsx")
    setup_logging(log_file_path)

    folder_structure = {
        "Accounts Receivable": [
            "Deleted Charges", "Accounts Receivable Aging", "AR Transactions",
            "Open Charges", "Payments"
        ],
        "Accounts Payable": [
            "Details of Expenses", "Detail of Expenses by Supplier Report",
            "Summary of Expenses by Account Report", "Summary of Expenses by Supplier Report"
        ],
        "Inventory/Parts": [
            "Stock Value Detail", "Returned Parts", "Core Detail"
        ],
        "Sales": [
            "Sales Summary", "Payment Summary By Pay Date", "Part Sales Detail",
            "Payment Detail By RO", "R.O. Sales Ledger", "Sales By Service Category",
            "Technician Sales Summary", "Taxes And Core Detail", "Part Sales Detail"
        ],
        "Rankings": [
            "Labor By Technician", "Sales by Service Writer"
        ]
    }

    # Get today's date and weekday
    now = datetime.now()
    today = now.day
    current_weekday = now.weekday()  # Monday is 0, Sunday is 6
    is_month_end = (now.replace(day=28) + timedelta(days=4)).day <= 3

    for k in range(2, row_count + 1):
        print(f"Processing row {k}")

        clientname = sheet.cell(row=k, column=1).value
        username = sheet.cell(row=k, column=2).value
        password = sheet.cell(row=k, column=3).value
        report = sheet.cell(row=k, column=4).value
        subreport = sheet.cell(row=k, column=5).value
        store = sheet.cell(row=k, column=6).value
        frequency = sheet.cell(row=k, column=7).value
        daily = sheet.cell(row=k, column=8).value
        weekly = sheet.cell(row=k, column=9).value
        monthly = sheet.cell(row=k, column=10).value

        if all(value is None for value in [clientname, username, password, report, subreport, frequency]):
            print("No more valid data found. Ending process.")
            break

        print(f"Clientname is: {clientname}")
        print(f"Username is: {username}")
        print(f"Password is: {password}")
        print(f"Report is: {report}")
        print(f"Subreport is: {subreport}")
        print(f"Store Name is: {store}")
        print(f"Frequency is: {frequency}")
        print(f"Daily: {daily}")
        print(f"Weekly: {weekly}")
        print(f"Monthly: {monthly}")

        try:
            # Daily Reports
            if daily == 'Yes':
                print("Processing Daily Report")
                process_report(driver, sheet, k, 'Daily', log_file_path, destination_dir, folder_structure, download_path)

            # Weekly Reports
            if weekly == 'Yes' and current_weekday == 0:  # 0 is Monday
                print("Processing Weekly Report")
                process_report(driver, sheet, k, 'Weekly', log_file_path, destination_dir, folder_structure, download_path)

            # Monthly Reports
            if monthly == 'Yes' and is_month_end:
                print("Processing Monthly Report")
                process_report(driver, sheet, k, 'Monthly', log_file_path, destination_dir, folder_structure, download_path)

        except Exception as e:
            print(f"An error occurred while processing row {k}: {e}")
            log_operation(log_file_path, clientname, username, password, report, subreport, frequency, "Fail", str(e))
            driver.refresh()

def process_report(driver, sheet, row, report_type, log_file_path, destination_dir, folder_structure, download_path):
    try:
        clientname = sheet.cell(row=row, column=1).value
        username = sheet.cell(row=row, column=2).value
        password = sheet.cell(row=row, column=3).value
        report = sheet.cell(row=row, column=4).value
        subreport = sheet.cell(row=row, column=5).value
        store = sheet.cell(row=row, column=6).value
        frequency = sheet.cell(row=row, column=7).value

        # Login and navigate to the report as before
        driver.get("https://cs.rowriter.com/")
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "loginLink"))).click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@name='Username']"))).send_keys(username)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@name='Password']"))).send_keys(password)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "/html/body/main/div/div[3]/div/div/form/div[4]/button"))).click()

        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "reports"))).click()

        report_selectors = {
            "Accounts Payable": "li:nth-of-type(1) > .sub-lvl",
            "Accounts Receivable": "li:nth-of-type(2) > ul > li:nth-of-type(2) > .sub-lvl",
            "Inventory/Parts": "li:nth-of-type(2) > ul > li:nth-of-type(3) > .sub-lvl",
            "Marketing": "li:nth-of-type(2) > ul > li:nth-of-type(4) > .sub-lvl",
            "Rankings": "li:nth-of-type(2) > ul > li:nth-of-type(5) > .sub-lvl",
            "Sales": ".sub-lvl[href='#Sales']"
        }

        if report.strip() in report_selectors:
            print(f"Clicking on {report} report...")
            WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, report_selectors[report.strip()]))).click()
            print("Clicked successfully!")
        else:
            print(f"Report '{report}' does not match expected values.")
            log_operation(log_file_path, clientname, username, password, report, subreport, frequency, "Fail", "Invalid report")
            return

        data_title = subreport.strip()
        elements = driver.find_elements(By.TAG_NAME, "a")
        for element in elements:
            if element.get_attribute("data-title") == data_title:
                element.click()
                break

        try:
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//select[@id='filter-shop-selector']")))
            dropdown = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//select[@id='filter-shop-selector']")))
            print("Clicking dropdown")
            dropdown.click()

            options = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.XPATH, "//optgroup/option"))
            )

            print(f"List size of shop dropdown: {len(options)}")
            link_texts = []
            for option in options:
                try:
                    link_texts.append(option.text)
                except Exception as e:
                    print(f"Error getting text from option: {e}")

            print(link_texts)

            data_title1 = store.strip()
            store_selected = False
            for option in options:
                if option.text.strip() == data_title1:
                    option.click()
                    print(f"Selected store: {store}")
                    store_selected = True
                    break

            if not store_selected:
                print(f"Store '{store}' not found in the dropdown.")
                log_operation(log_file_path, clientname, username, password, report, subreport, frequency, "Fail", f"Store '{store}' not found in dropdown")

        except TimeoutException:
            print("Element not found or not clickable: //select[@id='filter-shop-selector']")
            log_operation(log_file_path, clientname, username, password, report, subreport, frequency, "Fail", "Filter shop selector not found")

        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div[class='input-group input-daterange'] span[type='text']"))).click()

        frequency_ids = {
            "Today": "today_dateReport",
            "Yesterday": "yesterday_dateReport",
            "Last 7days": "last7days_dateReport",
            "Last 30 days": "last30days_dateReport",
            "This Month": "thisMonth_dateReport",
            "Last Month": "lastMonth_dateReport"
        }

        if frequency.strip() in frequency_ids:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, frequency_ids[frequency.strip()]))).click()
            print(f"Frequency selected: {frequency}")
        else:
            print(f"Frequency '{frequency}' does not match expected values.")
            log_operation(log_file_path, clientname, username, password, report, subreport, frequency, "Fail", "Invalid frequency")
            return

        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "(//button[@class='btn btn-primary rounded-0'][normalize-space()='Apply'])[1]"))).click()

        right = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/main/section/section/div[2]/div[4]/div/div[2]/div[2]/div/div[5]")))
        ActionChains(driver).move_to_element(right).context_click().perform()

        checkboxes = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, "//input[@type='checkbox']")))
        for checkbox in checkboxes:
            click_checkbox(driver, checkbox)

        export_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Export']")))
        export_button.click()
        print("Clicked Export button")

        excel_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Excel']")))
        excel_button.click()
        print("Clicked Excel button")

        if wait_for_downloads(download_path):
            list_of_files = os.listdir(download_path)
            full_path = [os.path.join(download_path, file) for file in list_of_files if not file.endswith('.crdownload')]
            if full_path:
                latest_file = max(full_path, key=os.path.getctime)
                now = datetime.now()
                year = now.strftime('%Y')
                month = now.strftime('%B')
                day = now.strftime('%d')
                today_date = now.strftime('%d%m%Y')
                current_time = now.strftime('%H%M%S')
                new_filename = f"{clientname}_{subreport}_{today_date}_{current_time}_RO_{report_type}.xlsx"
                new_file_path = os.path.join(download_path, new_filename)

                try:
                    os.rename(latest_file, new_file_path)
                    print(f"File renamed to {new_file_path}")

                    file_parts = [part.strip() for part in new_filename.split('_')]
                    if len(file_parts) >= 4:
                        clientname = file_parts[0]
                        report_name = file_parts[1]
                        date_folder = file_parts[2]

                        category_folder = None
                        for category, reports in folder_structure.items():
                            if report_name in reports:
                                category_folder = category
                                break

                        if category_folder:
                            # Create folders with year, month, and date
                            report_folder_path = create_folder_structure(destination_dir, clientname, year, month, day, category_folder, report_name)
                            destination_file = os.path.join(report_folder_path, new_filename)

                            try:
                                shutil.copy2(new_file_path, destination_file)
                                status = "Success"
                                print(f"Copied '{new_filename}' to '{destination_file}'.")

                            except (shutil.Error, OSError) as e:
                                status = "Fail"
                                print(f"Error copying '{new_filename}': {e}")

                            log_operation(log_file_path, clientname, username, password, report, subreport, frequency, status)

                        else:
                            print(f"No category found for report '{report_name}'. Skipping file '{new_filename}'.")
                    else:
                        print(f"Ignoring file '{new_filename}' as it does not match expected format.")

                except FileNotFoundError:
                    print(f"File {latest_file} not found for renaming, continuing process.")
            else:
                print("No files found in the Downloads directory.")
        else:
            print("File download timed out or failed.")

        # Logout logic
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[@id='login']"))).click()
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space()='Log-off']"))).click()

    except Exception as e:
        print(f"An error occurred: {e}")
        log_operation(log_file_path, clientname, username, password, report, subreport, frequency, "Fail", str(e))
        driver.refresh()

def main():
    driver, sheet, row_count, download_path = initialize_environment()
    process_rows(driver, sheet, row_count, download_path)
    driver.quit()

if __name__ == "__main__":
    main()
